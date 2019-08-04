import openpyxl

# wb = openxl.workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)

sheet = wb["Sheet1"]
"""
cell = sheet["a1"]
print(cell.row)
print(cell.column)
print(cell.coordinate)
"""

"""
cell = sheet.cell(row=1, column=1)

print(cell.row)
print(cell.column)
print(cell.coordinate)


print(sheet.max_row)
print(sheet.max_column)
"""


# iterate of rwo and columns. first row is 1
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        all = sheet.cell(row, col)
        print(all.value)


# sheet["a"]
# sheet["a:c""]
# sheet["a1:c4""]
# sheet["1:4""]

sheet.append([1,2,3])

wb.save("ok.xlsx")